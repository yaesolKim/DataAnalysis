import openpyxl

import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.ticker import MultipleLocator, FormatStrFormatter, MaxNLocator

class Data:

    def __init__(self, ws):
        self._num_row = ws.max_row
        self._time = np.array([ws.cell(row=i, column=1).value for i in range(2, self._num_row+1)])
        self._Ke = [ws.cell(row=i, column=12).value for i in range(2, self._num_row+1)]
        self._eth_pose = np.array([ws.cell(row=i, column=6).value for i in range(2, self._num_row+1)])
        self._rem_pose = np.array([ws.cell(row=i, column=7).value for i in range(2, self._num_row+1)])
        self._eth_force = np.array([ws.cell(row=i, column=9).value for i in range(2, self._num_row+1)])
        self._rem_force = np.array([ws.cell(row=i, column=10).value for i in range(2, self._num_row+1)])

def plot_motion(y1t, y1e, y1r, y2t, y2e, y2r, y3t, y3e, y3r, y4t, y4e, y4r):
    fig, axes = plt.subplots(nrows=2, ncols=2, figsize=(10, 8))

    axes[0, 0].plot(y1t, y1e, label='eth motion')
    axes[0, 0].plot(y1t, y1r, label='remote motion')
    axes[0,0].set_ylim(0, 0.25)
    axes[0, 0].legend()
    axes[0, 0].set_title('K = 20')

    axes[0, 1].plot(y2t, y2e, label='eth motion')
    axes[0, 1].plot(y2t, y2r, label='remote motion')
    axes[0, 1].set_ylim(0, 0.25)
    axes[0, 1].legend()
    axes[0, 1].set_title('K = 50')

    axes[1, 0].plot(y3t, y3e, label='eth motion')
    axes[1, 0].plot(y3t, y3r, label='remote motion')
    axes[1, 0].set_ylim(0, 0.25)
    axes[1, 0].legend()
    axes[1, 0].set_title('K = 60')

    axes[1, 1].plot(y4t, y4e, label='eth motion')
    axes[1, 1].plot(y4t, y4r, label='remote motion')
    axes[1, 1].set_ylim(0, 0.25)
    axes[1, 1].legend()
    axes[1, 1].set_title('K = Ke')

    plt.tight_layout()
    plt.show()

def plot_force(y1t, y1e, y1r, y2t, y2e, y2r, y3t, y3e, y3r, y4t, y4e, y4r):
    fig, axes = plt.subplots(nrows=2, ncols=2, figsize=(10, 8))

    axes[0, 0].plot(y1t, y1e, label='eth force')
    axes[0, 0].plot(y1t, y1r, label='remote force')
    axes[0, 0].set_ylim(0, 20)
    axes[0, 0].legend()
    axes[0, 0].set_title('K = 20')

    axes[0, 1].plot(y2t, y2e, label='eth force')
    axes[0, 1].plot(y2t, y2r, label='remote force')
    axes[0, 1].set_ylim(0, 20)
    axes[0, 1].legend()
    axes[0, 1].set_title('K = 50')

    axes[1, 0].plot(y3t, y3e, label='eth force')
    axes[1, 0].plot(y3t, y3r, label='remote force')
    axes[1, 0].set_ylim(0, 20)
    axes[1, 0].legend()
    axes[1, 0].set_title('K = 60')

    axes[1, 1].plot(y4t, y4e, label='eth force')
    axes[1, 1].plot(y4t, y4r, label='remote force')
    axes[1, 1].set_ylim(0, 20)
    axes[1, 1].legend()
    axes[1, 1].set_title('K = Ke')

    plt.tight_layout()
    plt.show()

def plot_Ke(time, Ke):
    # Plot the data
    plt.plot(time, Ke, label='')
    plt.xlabel('Time (s)')
    plt.ylabel('Ke (Nm)')
    plt.title('Variable stiffness profile over time')

    # Print the RMS of motion jerk on the plot
    plt.text(0.5, np.max(Ke), f'Mean of Ke: {np.mean(Ke):.2f}', ha='center', va='bottom')

    # Show the plot
    plt.legend()
    plt.grid(True)
    plt.show()

def plot_diff(data, t1, diff_1, t2, diff_2, t3, diff_3, t4, diff_4):
    fig, axes = plt.subplots(nrows=2, ncols=2, figsize=(10, 8))

    if data == "motion":
        plt.title('Motion Difference between ETH and Remote robot', fontsize=16)
        lim = (-0.15, 0.15)
        text_loc_x = 0.0
        text_loc_y = -0.1

    elif data == "force":
        plt.title('Force Difference between ETH and Remote robot', fontsize=16)
        ylim = (-10, 15)
        text_loc_x = 0.0
        text_loc_y = 10

    axes[0, 0].plot(t1, diff_1)
    axes[0, 0].set_ylim(lim)
    axes[0, 0].set_title('K = 20')
    rmse_1 = str(np.sqrt(np.mean(diff_1 ** 2)))
    axes[0, 0].text(text_loc_x, text_loc_y, 'RMSE = ' + rmse_1, fontsize=12, bbox=dict(facecolor='white', alpha=0.5))

    axes[0, 1].plot(t2, diff_2)
    axes[0, 1].set_ylim(lim)
    axes[0, 1].set_title('K = 50')
    rmse_2 = str(np.sqrt(np.mean(diff_2 ** 2)))
    axes[0, 1].text(text_loc_x, text_loc_y, 'RMSE = ' + rmse_2, fontsize=12, bbox=dict(facecolor='white', alpha=0.5))

    axes[1, 0].plot(t3, diff_3)
    axes[1, 0].set_ylim(lim)
    axes[1, 0].set_title('K = 60')
    rmse_3 = str(np.sqrt(np.mean(diff_3 ** 2)))
    axes[1, 0].text(text_loc_x, text_loc_y, 'RMSE = ' + rmse_3, fontsize=12, bbox=dict(facecolor='white', alpha=0.5))

    axes[1, 1].plot(t4, diff_4)
    axes[1, 1].set_ylim(lim)
    axes[1, 1].set_title('K = Ke')
    rmse_4 = str(np.sqrt(np.mean(diff_4 ** 2)))
    axes[1, 1].text(text_loc_x, text_loc_y, 'RMSE = ' + rmse_4, fontsize=12, bbox=dict(facecolor='white', alpha=0.5))

    # Add spacing between subplots
    plt.tight_layout()
    plt.show()


def bar_plot(data, y1, y2, y3, y4):
    y = {'20': y1, '50': y2, '60': y3, 'Ke': y4}
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.bar(y.keys(), y.values(), width=0.4, edgecolor='black', color=['green', 'red', 'orange', 'blue'])

    for i in ax.patches:
        plt.text(i.get_x() + 0.1, i.get_height() + 0.1, str(round((i.get_height()), 2)), fontsize=10, fontweight='bold', color='grey')

    if data == "time":
        plt.xlabel("Stiffness Condition (Nm)")
        plt.ylabel("Task Completion Time (sec)")

    elif data == "peak_eth_force":
        plt.xlabel("ETH Peak Force (N)")
        plt.ylabel("Task Completion Time (sec)")

    elif data == "peak_rem_force":
        plt.xlabel("Remote Peak Force (N)")
        plt.ylabel("Task Completion Time (sec)")

    elif data == "total_eth_force":
        plt.xlabel("ETH Total Force (N)")
        plt.ylabel("Task Completion Time (sec)")

    elif data == "total_rem_force":
        plt.xlabel("Remote Total Force (N)")
        plt.ylabel("Task Completion Time (sec)")

    plt.show()


wb_name = 'xlsx/data.xlsx'
ws1_name = '20 (2)'
ws2_name = '50 (2)'
ws3_name = '60 (2)'
ws4_name = 'Ke (2)'

wb = openpyxl.load_workbook(wb_name)
Data_20 = Data(wb[ws1_name])
Data_50 = Data(wb[ws2_name])
Data_60 = Data(wb[ws3_name])
Data_Ke = Data(wb[ws4_name])

# set global parameters
mpl.rcParams['pdf.fonttype'] = 42
mpl.rcParams['ps.fonttype'] = 42
mpl.rcParams['font.family'] = 'Arial'

# draw motion profile of each conditions
#plot_motion(Data_20._time, Data_20._eth_pose, Data_20._rem_pose, Data_50._time, Data_50._eth_pose, Data_50._rem_pose, Data_60._time, Data_60._eth_pose, Data_60._rem_pose, Data_Ke._time, Data_Ke._eth_pose, Data_Ke._rem_pose)
#plot_force(Data_20._time, Data_20._eth_force, Data_20._rem_force, Data_50._time, Data_50._eth_force, Data_50._rem_force, Data_60._time, Data_60._eth_force, Data_60._rem_force, Data_Ke._time, Data_Ke._eth_force, Data_Ke._rem_force)

## motion difference and force difference between eth robot and remote robot
#plot_diff("motion", Data_20._time, Data_20._eth_pose-Data_20._rem_pose, Data_50._time, Data_50._eth_pose-Data_50._rem_pose, Data_60._time, Data_60._eth_pose-Data_60._rem_pose, Data_Ke._time, Data_Ke._eth_pose-Data_Ke._rem_pose)
#plot_diff("force", Data_20._time, Data_20._eth_force-Data_20._rem_force, Data_50._time, Data_50._eth_force-Data_50._rem_force, Data_60._time, Data_60._eth_force-Data_60._rem_force, Data_Ke._time, Data_Ke._eth_force-Data_Ke._rem_force)

## Ke profile
#plot_Ke(Data_Ke._time, Data_Ke._Ke)

## task completion time
#bar_plot("time", Data_20._time[-1], Data_50._time[-1], Data_60._time[-1], Data_Ke._time[-1])

## peak force
#bar_plot("peak_eth_force", np.max(Data_20._eth_force), np.max(Data_50._eth_force), np.max(Data_60._eth_force), np.max(Data_Ke._eth_force))
#bar_plot("peak_rem_force", np.max(Data_20._rem_force), np.max(Data_50._rem_force), np.max(Data_60._rem_force), np.max(Data_Ke._rem_force))
#
## total force
#bar_plot("total_eth_force", np.sum(Data_20._eth_force), np.sum(Data_50._eth_force), np.sum(Data_60._eth_force), np.sum(Data_Ke._eth_force))
#bar_plot("total_rem_force", np.sum(Data_20._rem_force), np.sum(Data_50._rem_force), np.sum(Data_60._rem_force), np.sum(Data_Ke._rem_force))
