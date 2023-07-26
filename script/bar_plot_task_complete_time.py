import openpyxl

import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.ticker import MultipleLocator, FormatStrFormatter, MaxNLocator

class Data:

    def __init__(self, ws):
        self._num_row = ws.max_row
        self._time = [ws.cell(row=i, column=1).value for i in range(2, self._num_row+1)]

wb_name = 'xlsx/data.xlsx'
ws1_name = '20 (2)'
ws2_name = '50 (2)'
ws3_name = '60 (2)'
ws4_name = 'Ke (2)'

wb = openpyxl.load_workbook(wb_name)
Data_20 = Data(wb[ws1_name])
Data_50 = Data(wb[ws2_name])
Data_60 = Data(wb[ws3_name])
Data_Ke = Data(wb[ws4_name])

y1 = Data_20._time[Data_20._num_row-2]
y2 = Data_50._time[Data_50._num_row-2]
y3 = Data_60._time[Data_60._num_row-2]
y4 = Data_Ke._time[Data_Ke._num_row-2]

# set global parameters
mpl.rcParams['pdf.fonttype'] = 42
mpl.rcParams['ps.fonttype'] = 42
mpl.rcParams['font.family'] = 'Arial'

y = {'20':y1, '50':y2, '60':y3, 'Ke':y4}

fig, ax = plt.subplots()
ax.bar(y.keys(), y.values(), width=0.4, edgecolor='black', color=['green', 'red', 'orange', 'blue'])

for i in ax.patches:
    plt.text(i.get_x()+0.1, i.get_height() + 0.1, str(round((i.get_height()), 2)),
             fontsize=10, fontweight='bold', color='grey')

plt.xlabel("Stiffness Condition (Nm)")
plt.ylabel("Task Completion Time (sec)")

plt.show()
