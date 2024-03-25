import openpyxl

import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.ticker import MultipleLocator, FormatStrFormatter, MaxNLocator

class Data:

    def __init__(self, ws):
        self._num_row = ws.max_row

        self._start = 0.00
        self._finish = 0.00

        #self._eth_pose = [ws.cell(row=i, column=6).value for i in range(2, self._num_row+1)]
        self._rem_pose = [ws.cell(row=i, column=7).value for i in range(2, self._num_row+1)]
        #self._eth_force = [ws.cell(row=i, column=9).value for i in range(2, self._num_row+1)]
        #self._rem_force = [ws.cell(row=i, column=10).value for i in range(2, self._num_row+1)]

    def find_start_finish(self):
        for index, number in enumerate(self._rem_pose):
            if number >= 0.625 and self._start == 0.00:
                self._start = index
            if number > 0.715:
                self._finish = index
                return 1
        return 0

def delete_rows_in_worksheet(filename, worksheet_name, start_row, end_row):
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook[worksheet_name]

    worksheet.delete_rows(start_row, end_row)
    workbook.save(filename)
    print(f"Rows {start_row} to {end_row} deleted from worksheet '{worksheet_name}' in '{filename}'.")


wb_name = 'ETH_VIC/data.ETH_VIC'
ws1_name = '20 (2)'
ws2_name = '50 (2)'
ws3_name = '60 (2)'
ws4_name = 'Ke (2)'

wb = openpyxl.load_workbook(wb_name)

Data_20 = Data(wb[ws1_name])
Data_50 = Data(wb[ws2_name])
Data_60 = Data(wb[ws3_name])
Data_Ke = Data(wb[ws4_name])

Data_20.find_start_finish()
Data_50.find_start_finish()
Data_60.find_start_finish()
Data_Ke.find_start_finish()

delete_rows_in_worksheet(wb_name, ws1_name, Data_20._finish, Data_20._num_row)
delete_rows_in_worksheet(wb_name, ws1_name, 2, Data_20._start)
delete_rows_in_worksheet(wb_name, ws2_name, Data_50._finish, Data_50._num_row)
delete_rows_in_worksheet(wb_name, ws2_name, 2, Data_50._start)
delete_rows_in_worksheet(wb_name, ws3_name, Data_60._finish, Data_60._num_row)
delete_rows_in_worksheet(wb_name, ws3_name, 2, Data_60._start)
delete_rows_in_worksheet(wb_name, ws4_name, Data_Ke._finish, Data_Ke._num_row)
delete_rows_in_worksheet(wb_name, ws4_name, 2, Data_Ke._start)
