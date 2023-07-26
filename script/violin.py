import openpyxl

wb = openpyxl.load_workbook('S1.xlsx')
ws = wb['Sheet1']
#print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))

start = ws.cell(row=1, column=1).value.split('/')
start_time = start[3].split(':') #hour/min/sec

finish = ws.cell(row=ws.max_row, column=1).value.split('/')
finish_time = finish[3].split(':') #hour/min/sec

#print(start_time)
#print(finish_time)
t_i = [float(finish_time[i]) - float(start_time[i]) for i in range (0,3)]

#print("finish time: " + str(t_i))

#calculate in second
interaction_time = t_i[0]*60*60 + t_i[1]*60 + t_i[2]
print(interaction_time)