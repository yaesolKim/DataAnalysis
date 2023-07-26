import openpyxl

import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.ticker import MultipleLocator, FormatStrFormatter, MaxNLocator

class Data:

    def __init__(self, ws):
        self._num_row = ws.max_row
        self._time_start = ws.cell(row=2, column=1).value.split('/')[3].split(':')
        self._time = [ws.cell(row=i, column=1).value.split('/')[3] for i in range(2, self._num_row+1)]

    def set_time_start_from_zero(self):
        for i in range(0, self._num_row - 1):
            h = float(self._time[i].split(':')[0]) - float(self._time_start[0])
            m = float(self._time[i].split(':')[1]) - float(self._time_start[1])
            s = float(self._time[i].split(':')[2]) - float(self._time_start[2])
            self._time[i] = h*60*60 + m*60 + s

    def interaction_time_sec(self):
        if self.is_time_from_zero() is None:
            self.set_time_start_from_zero()
        return self._time[self._num_row-2]

    def is_time_from_zero(self):
        try:
            start_time = float(self._time[0])
            return start_time
        except ValueError as ve:
            return None
        except AttributeError as ae:
            return None

wb = openpyxl.load_workbook('xlsx/4_all_at_once.xlsx')

Data_20 = Data(wb['20 (2)'])
Data_50 = Data(wb['50 (2)'])
Data_60 = Data(wb['60 (2)'])
Data_Ke = Data(wb['Ke (2)'])

y1= Data_20.interaction_time_sec()
y2= Data_50.interaction_time_sec()
y3= Data_60.interaction_time_sec()
y4= Data_Ke.interaction_time_sec()

# set global parameters
mpl.rcParams['pdf.fonttype'] = 42
mpl.rcParams['ps.fonttype'] = 42
mpl.rcParams['font.family'] = 'Arial'

y = {'20':y1, '50':y2, '60':y3, 'Ke':y4}

fig,ax = plt.subplots()
ax.bar(y.keys(), y.values(), width=0.4, edgecolor='black', color=['green','red','orange','blue'])

for i in ax.patches:
    plt.text(i.get_x()+0.1, i.get_height() + 0.1, str(round((i.get_height()), 2)),
             fontsize=10, fontweight='bold', color='grey')

plt.xlabel("Stiffness Condition (Nm)")
plt.ylabel("Task Completion Time (sec)")

plt.show()
