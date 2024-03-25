import openpyxl

import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.ticker import MultipleLocator, FormatStrFormatter, MaxNLocator

class Data:

    def __init__(self, ws):
        self._num_row = ws.max_row

        self._eth_pose_start = ws.cell(row=2, column=6).value
        self._rem_pose_start = ws.cell(row=2, column=7).value
        self._eth_force_start = ws.cell(row=2, column=9).value
        self._rem_force_start = ws.cell(row=2, column=10).value

        self._eth_pose = [ws.cell(row=i, column=6).value for i in range(2, self._num_row+1)]
        self._rem_pose = [ws.cell(row=i, column=7).value for i in range(2, self._num_row+1)]
        self._eth_force = [ws.cell(row=i, column=9).value for i in range(2, self._num_row+1)]
        self._rem_force = [ws.cell(row=i, column=10).value for i in range(2, self._num_row+1)]

    def set_motion_from_start(self):
        self._eth_pose = [self._eth_pose_start - self._eth_pose[i] for i in range(0, self._num_row - 1)]
        self._rem_pose = [self._rem_pose[i] - self._rem_pose_start for i in range(0, self._num_row - 1)]

    def set_force_from_start(self):
        self._eth_force = [self._eth_force[i] - self._eth_force_start for i in range(0, self._num_row - 1)]
        self._rem_force = [self._rem_force[i] - self._rem_force_start for i in range(0, self._num_row - 1)]

def change_data_in_worksheet(filename, worksheet_name, num_row, ep, ef, rp, rf):
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook[worksheet_name]

    # Change the value in the specified cell
    for i in range(2, num_row+1):
        worksheet.cell(row=i, column=6, value=ep[i-2])
        worksheet.cell(row=i, column=7, value=ef[i-2])
        worksheet.cell(row=i, column=9, value=rp[i-2])
        worksheet.cell(row=i, column=10, value=rf[i-2])

    # Save the modified workbook
    workbook.save(filename)
    print(f"Cells updated in worksheet '{worksheet_name}' in '{filename}'.")

wb_name = 'ETH_VIC/data.ETH_VIC'
ws1_name = '20 (2)'
ws2_name = '50 (2)'
ws3_name = '60 (2)'
ws4_name = 'Ke (2)'

wb = openpyxl.load_workbook(wb_name)

Data_20 = Data(wb[ws1_name])
Data_50 = Data(wb[ws2_name])
Data_60 = Data(wb[ws3_name])
Data_Ke = Data(wb[ws4_name])

Data_20.set_motion_from_start()
Data_20.set_force_from_start()
Data_50.set_motion_from_start()
Data_50.set_force_from_start()
Data_60.set_motion_from_start()
Data_60.set_force_from_start()
Data_Ke.set_motion_from_start()
Data_Ke.set_force_from_start()

change_data_in_worksheet(wb_name, ws1_name, Data_20._num_row, Data_20._eth_pose, Data_20._rem_pose, Data_20._eth_force ,Data_20._rem_force)
change_data_in_worksheet(wb_name, ws2_name, Data_50._num_row, Data_50._eth_pose, Data_50._rem_pose, Data_50._eth_force ,Data_50._rem_force)
change_data_in_worksheet(wb_name, ws3_name, Data_60._num_row, Data_60._eth_pose, Data_60._rem_pose, Data_60._eth_force ,Data_60._rem_force)
change_data_in_worksheet(wb_name, ws4_name, Data_Ke._num_row, Data_Ke._eth_pose, Data_Ke._rem_pose, Data_Ke._eth_force ,Data_Ke._rem_force)
