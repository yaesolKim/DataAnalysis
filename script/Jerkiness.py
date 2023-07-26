import openpyxl

import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.ticker import MultipleLocator, FormatStrFormatter, MaxNLocator

class Data:

    def __init__(self, ws):
        self._num_row = ws.max_row

        self._time_start = ws.cell(row=2, column=1).value.split('/')[3].split(':')
        self._time = [ws.cell(row=i, column=1).value.split('/')[3] for i in range(2, self._num_row+1)]

        self._eth_pose_start = ws.cell(row=2, column=6).value
        self._rem_pose_start = ws.cell(row=2, column=7).value

        self._eth_force_start = ws.cell(row=2, column=9).value
        self._rem_force_start = ws.cell(row=2, column=10).value

        self._eth_pose = [ws.cell(row=i, column=6).value for i in range(2, self._num_row+1)]
        self._rem_pose = [ws.cell(row=i, column=7).value for i in range(2, self._num_row+1)]
        self._eth_force = [ws.cell(row=i, column=9).value for i in range(2, self._num_row+1)]
        self._rem_force = [ws.cell(row=i, column=10).value for i in range(2, self._num_row+1)]

        self._eth_pose_jerk = [0.0 for i in range(0, self._num_row - 1)]
        self._rem_pose_jerk = [0.0 for i in range(0, self._num_row - 1)]
        self._eth_force_jerk = [0.0 for i in range(0, self._num_row - 1)]
        self._rem_force_jerk = [0.0 for i in range(0, self._num_row - 1)]

        self._eth_motion_jerk_rms = 0.00
        self._rem_motion_jerk_rms = 0.00
        self._eth_force_jerk_rms = 0.00
        self._rem_force_jerk_rms = 0.00



    def set_time_start_from_zero(self):
        for i in range(0, self._num_row - 1):
            h = float(self._time[i].split(':')[0]) - float(self._time_start[0])
            m = float(self._time[i].split(':')[1]) - float(self._time_start[1])
            s = float(self._time[i].split(':')[2]) - float(self._time_start[2])
            self._time[i] = h*60*60 + m*60 + s

    def is_time_from_zero(self):
        try:
            start_time = float(self._time[0])
            return start_time
        except ValueError as ve:
            return None
        except AttributeError as ae:
            return None

    def interaction_time_sec(self):
        if self.is_time_from_zero() is None:
            self.set_time_start_from_zero()
        return self._time[self._num_row-2]

    def is_motion_from_zero(self):
        return float(self._eth_pose[0])

    def set_motion_from_start(self):
        self._eth_pose = [self._eth_pose_start - self._eth_pose[i] for i in range(0, self._num_row - 1)]
        self._rem_pose = [self._rem_pose[i] - self._rem_pose_start for i in range(0, self._num_row - 1)]

    def set_force_from_start(self):
        self._eth_force = [self._eth_force[i] - self._eth_force_start for i in range(0, self._num_row - 1)]
        self._rem_force = [self._rem_force[i] - self._rem_force_start for i in range(0, self._num_row - 1)]

    def preprocessing(self):
        self.set_force_from_start()
        self.set_motion_from_start()
        self.set_time_start_from_zero()



    def motion_jerk(self):
        # Calculate velocity using finite differences
        velocity = np.gradient(self._eth_pose, self._time, edge_order=2)
        acceleration = np.gradient(velocity, self._time, edge_order=2)
        jerk = np.gradient(acceleration, self._time, edge_order=2)
        self._eth_pose_jerk = [abs(jerk[i]) for i in range(0, self._num_row - 1)]

        velocity = np.gradient(self._rem_pose, self._time, edge_order=2)
        acceleration = np.gradient(velocity, self._time, edge_order=2)
        jerk = np.gradient(acceleration, self._time, edge_order=2)
        self._rem_pose_jerk = [abs(jerk[i]) for i in range(0, self._num_row - 1)]

        self._eth_motion_jerk_rms = np.sqrt(np.mean(self._eth_pose_jerk) ** 2)
        self._rem_motion_jerk_rms = np.sqrt(np.mean(self._rem_pose_jerk) ** 2)


    def force_jerk(self):
        velocity = np.gradient(self._eth_force, self._time, edge_order=2)
        acceleration = np.gradient(velocity, self._time, edge_order=2)
        jerk = np.gradient(acceleration, self._time, edge_order=2)
        self._eth_force_jerk = [abs(jerk[i]) for i in range(0, self._num_row - 1)]

        velocity = np.gradient(self._rem_force, self._time, edge_order=2)
        acceleration = np.gradient(velocity, self._time, edge_order=2)
        jerk = np.gradient(acceleration, self._time, edge_order=2)
        self._rem_force_jerk = [abs(jerk[i]) for i in range(0, self._num_row - 1)]

        self._eth_force_jerk_rms = np.sqrt(np.mean(self._eth_force_jerk) ** 2)
        self._rem_force_jerk_rms = np.sqrt(np.mean(self._rem_force_jerk) ** 2)



def plot_jerk(y1, y2, y3, y4, robot, profile):
    fig, axes = plt.subplots(nrows=2, ncols=2, figsize=(10, 6))

    if robot == "eth":
        if profile == "pose":
            y1_y = y1._eth_pose_jerk
            y2_y = y2._eth_pose_jerk
            y3_y = y3._eth_pose_jerk
            y4_y = y4._eth_pose_jerk
            y1_rms = y1._eth_motion_jerk_rms
            y2_rms = y2._eth_motion_jerk_rms
            y3_rms = y3._eth_motion_jerk_rms
            y4_rms = y4._eth_motion_jerk_rms
            y_axis = 6000

        elif profile == "force":
            y1_y = y1._eth_force_jerk
            y2_y = y2._eth_force_jerk
            y3_y = y3._eth_force_jerk
            y4_y = y4._eth_force_jerk
            y1_rms = y1._eth_force_jerk_rms
            y2_rms = y2._eth_force_jerk_rms
            y3_rms = y3._eth_force_jerk_rms
            y4_rms = y4._eth_force_jerk_rms
            y_axis = 700000

    elif robot == "rem":
        if profile == "pose":
            y1_y = y1._rem_pose_jerk
            y2_y = y2._rem_pose_jerk
            y3_y = y3._rem_pose_jerk
            y4_y = y4._rem_pose_jerk
            y1_rms = y1._rem_motion_jerk_rms
            y2_rms = y2._rem_motion_jerk_rms
            y3_rms = y3._rem_motion_jerk_rms
            y4_rms = y4._rem_motion_jerk_rms
            y_axis = 2000

        elif profile == "force":
            y1_y = y1._rem_force_jerk
            y2_y = y2._rem_force_jerk
            y3_y = y3._rem_force_jerk
            y4_y = y4._rem_force_jerk
            y1_rms = y1._rem_force_jerk_rms
            y2_rms = y2._rem_force_jerk_rms
            y3_rms = y3._rem_force_jerk_rms
            y4_rms = y4._rem_force_jerk_rms
            y_axis = 500000

    axes[0, 0].plot(y1._time, y1_y, label='')
    axes[0, 0].set_ylim(0, y_axis)
    axes[0, 0].set_title('K = 20')
    axes[0, 0].text(np.max(y1._time)*0.2, y_axis*0.9, f'Mean Jerk RMS: {y1_rms:.2f}', ha='center', va='bottom')

    axes[0, 1].plot(y2._time, y2_y, label='')
    axes[0, 1].set_ylim(0, y_axis)
    axes[0, 1].set_title('K = 50')
    axes[0, 1].text(np.max(y2._time)*0.2, y_axis*0.9, f'Mean Jerk RMS: {y2_rms:.2f}', ha='center', va='bottom')
    axes[1, 0].plot(y3._time, y3_y, label='')
    axes[1, 0].set_ylim(0, y_axis)
    axes[1, 0].set_title('K = 60')
    axes[1, 0].text(np.max(y3._time)*0.2, y_axis*0.9, f'Mean Jerk RMS: {y3_rms:.2f}', ha='center', va='bottom')
    axes[1, 1].plot(y4._time, y4_y, label='')
    axes[1, 1].set_ylim(0, y_axis)
    axes[1, 1].set_title('K = Ke')
    axes[1, 1].text(np.max(y4._time)*0.2, y_axis*0.9, f'Mean Jerk RMS: {y4_rms:.2f}', ha='center', va='bottom')

    # Add spacing between subplots
    plt.tight_layout()

    # Display the subplots
    plt.show()

def bar_plot_time(y1, y2, y3, y4):
    y = {'20': y1, '50': y2, '60': y3, 'Ke': y4}
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.bar(y.keys(), y.values(), width=0.4, edgecolor='black', color=['green', 'red', 'orange', 'blue'])

    for i in ax.patches:
        plt.text(i.get_x() + 0.1, i.get_height() + 0.1, str(round((i.get_height()), 2)), fontsize=10, fontweight='bold', color='grey')

    plt.xlabel("Stiffness Condition (Nm)")
    plt.ylabel("Task Completion Time (sec)")

    plt.show()

def plot_motion_raw(y1, y2, y3, y4):
    fig, axes = plt.subplots(nrows=2, ncols=2, figsize=(10, 6))

    axes[0, 0].plot(y1._time, y1._eth_pose, label='eth motion')
    axes[0, 0].plot(y1._time, y1._rem_pose, label='remote motion')  # , marker='o', markersize=2)
    axes[0, 0].legend()
    axes[0, 0].set_title('K = 20')

    axes[0, 1].plot(y2._time, y2._eth_pose, label='eth motion')
    axes[0, 1].plot(y2._time, y2._rem_pose, label='remote motion')  # , marker='o', markersize=2)
    axes[0, 1].legend()
    axes[0, 1].set_title('K = 50')

    axes[1, 0].plot(y3._time, y3._eth_pose, label='eth motion')
    axes[1, 0].plot(y3._time, y3._rem_pose, label='remote motion')  # , marker='o', markersize=2)
    axes[1, 0].legend()
    axes[1, 0].set_title('K = 60')

    axes[1, 1].plot(y4._time, y4._eth_pose, label='eth motion')
    axes[1, 1].plot(y4._time, y4._rem_pose, label='remote motion')  # , marker='o', markersize=2)
    axes[1, 1].legend()
    axes[1, 1].set_title('K = Ke')

    # Add spacing between subplots
    plt.tight_layout()

    # Display the subplots
    plt.show()

def plot_motion(y1t, y1e, y1r, y2t, y2e, y2r, y3t, y3e, y3r, y4t, y4e, y4r):
    fig, axes = plt.subplots(nrows=2, ncols=2, figsize=(10, 8))

    axes[0, 0].plot(y1t, y1e, label='eth motion')
    axes[0, 0].plot(y1t, y1r, label='remote motion')  # , marker='o', markersize=2)
    axes[0,0].set_ylim(0, 0.25)
    axes[0, 0].legend()
    axes[0, 0].set_title('K = 20')

    axes[0, 1].plot(y2t, y2e, label='eth motion')
    axes[0, 1].plot(y2t, y2r, label='remote motion')  # , marker='o', markersize=2)
    axes[0, 1].set_ylim(0, 0.25)
    axes[0, 1].legend()
    axes[0, 1].set_title('K = 50')

    axes[1, 0].plot(y3t, y3e, label='eth motion')
    axes[1, 0].plot(y3t, y3r, label='remote motion')  # , marker='o', markersize=2)
    axes[1, 0].set_ylim(0, 0.25)
    axes[1, 0].legend()
    axes[1, 0].set_title('K = 60')

    axes[1, 1].plot(y4t, y4e, label='eth motion')
    axes[1, 1].plot(y4t, y4r, label='remote motion')  # , marker='o', markersize=2)
    axes[1, 1].set_ylim(0, 0.25)
    axes[1, 1].legend()
    axes[1, 1].set_title('K = Ke')

    # Add spacing between subplots
    plt.tight_layout()

    # Display the subplots
    plt.show()

def plot_force(y1t, y1e, y1r, y2t, y2e, y2r, y3t, y3e, y3r, y4t, y4e, y4r):
    fig, axes = plt.subplots(nrows=2, ncols=2, figsize=(10, 8))

    axes[0, 0].plot(y1t, y1e, label='eth force')
    axes[0, 0].plot(y1t, y1r, label='remote force')  # , marker='o', markersize=2)
    axes[0, 0].set_ylim(0, 20)
    axes[0, 0].legend()
    axes[0, 0].set_title('K = 20')

    axes[0, 1].plot(y2t, y2e, label='eth force')
    axes[0, 1].plot(y2t, y2r, label='remote force')  # , marker='o', markersize=2)
    axes[0, 1].set_ylim(0, 20)
    axes[0, 1].legend()
    axes[0, 1].set_title('K = 50')

    axes[1, 0].plot(y3t, y3e, label='eth force')
    axes[1, 0].plot(y3t, y3r, label='remote force')  # , marker='o', markersize=2)
    axes[1, 0].set_ylim(0, 20)
    axes[1, 0].legend()
    axes[1, 0].set_title('K = 60')

    axes[1, 1].plot(y4t, y4e, label='eth force')
    axes[1, 1].plot(y4t, y4r, label='remote force')  # , marker='o', markersize=2)
    axes[1, 1].set_ylim(0, 20)
    axes[1, 1].legend()
    axes[1, 1].set_title('K = Ke')

    # Add spacing between subplots
    plt.tight_layout()

    # Display the subplots
    plt.show()

wb = openpyxl.load_workbook('xlsx/4_all_at_once.xlsx')

Data_20 = Data(wb['20 (2)'])
Data_50 = Data(wb['50 (2)'])
Data_60 = Data(wb['60 (2)'])
Data_Ke = Data(wb['Ke (2)'])

Data_20.preprocessing()
Data_50.preprocessing()
Data_60.preprocessing()
Data_Ke.preprocessing()

# set global parameters
mpl.rcParams['pdf.fonttype'] = 42
mpl.rcParams['ps.fonttype'] = 42
mpl.rcParams['font.family'] = 'Arial'

Data_20.motion_jerk()
Data_50.motion_jerk()
Data_60.motion_jerk()
Data_Ke.motion_jerk()

Data_20.force_jerk()
Data_50.force_jerk()
Data_60.force_jerk()
Data_Ke.force_jerk()

plot_jerk(Data_20, Data_50, Data_60, Data_Ke, 'eth', 'pose')
plot_jerk(Data_20, Data_50, Data_60, Data_Ke, 'rem', 'pose')
plot_jerk(Data_20, Data_50, Data_60, Data_Ke, 'eth', 'force')
plot_jerk(Data_20, Data_50, Data_60, Data_Ke, 'rem', 'force')


#plot_jerk(Data_20._time, Data_20._eth_pose_jerk, Data_50._time, Data_50._eth_pose_jerk, Data_60._time, Data_60._eth_pose_jerk, Data_Ke._time, Data_Ke._eth_pose_jerk)
#plot_jerk(Data_20._jerk_time, Data_20._eth_pose_jerk, Data_50._jerk_time, Data_50._eth_pose_jerk, Data_60._jerk_time, Data_60._eth_pose_jerk, Data_Ke._jerk_time, Data_Ke._eth_pose_jerk)
#plot_jerk(Data_20._jerk_time, Data_20._rem_pose_jerk, Data_50._jerk_time, Data_50._rem_pose_jerk, Data_60._jerk_time, Data_60._rem_pose_jerk, Data_Ke._jerk_time, Data_Ke._rem_pose_jerk)
#plot_jerk(Data_20._jerk_time, Data_20._eth_force_jerk, Data_50._jerk_time, Data_50._eth_force_jerk, Data_60._jerk_time, Data_60._eth_force_jerk, Data_Ke._jerk_time, Data_Ke._eth_force_jerk)
#plot_jerk(Data_20._jerk_time, Data_20._rem_force_jerk, Data_50._jerk_time, Data_50._rem_force_jerk, Data_60._jerk_time, Data_60._rem_force_jerk, Data_Ke._jerk_time, Data_Ke._rem_force_jerk)

#draw barplot of task completion time
#bar_plot_time(Data_20.interaction_time_sec(), Data_50.interaction_time_sec(), Data_60.interaction_time_sec(), Data_Ke.interaction_time_sec())

# draw motion profile of each conditions
#plot_motion_raw(Data_20, Data_50, Data_60, Data_Ke) # raw
#plot_motion(Data_20._time, Data_20._eth_pose, Data_20._rem_pose, Data_50._time, Data_50._eth_pose, Data_50._rem_pose, Data_60._time, Data_60._eth_pose, Data_60._rem_pose, Data_Ke._time, Data_Ke._eth_pose, Data_Ke._rem_pose)
#plot_force(Data_20._time, Data_20._eth_force, Data_20._rem_force, Data_50._time, Data_50._eth_force, Data_50._rem_force, Data_60._time, Data_60._eth_force, Data_60._rem_force, Data_Ke._time, Data_Ke._eth_force, Data_Ke._rem_force)