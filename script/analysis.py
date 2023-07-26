import openpyxl

import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.ticker import MultipleLocator, FormatStrFormatter, MaxNLocator

class Data:

    def __init__(self, ws):
        self._num_row = ws.max_row

        self._time_start = ws.cell(row=2, column=1).value.split('/')[3].split(':')

        self._eth_pose_start = ws.cell(row=2, column=6).value
        self._rem_pose_start = ws.cell(row=2, column=7).value
        self._eth_force_start = ws.cell(row=2, column=9).value
        self._rem_force_start = ws.cell(row=2, column=10).value

        self._time_hour = [ws.cell(row=i, column=1).value.split('/')[3].split(':')[0] for i in range(2, self._num_row+1)]
        self._time_min = [ws.cell(row=i, column=1).value.split('/')[3].split(':')[1] for i in range(2, self._num_row+1)]
        self._time_sec = [ws.cell(row=i, column=1).value.split('/')[3].split(':')[2] for i in range(2, self._num_row+1)]
        self._time = [0.0 for i in range(2, self._num_row+1)]

        self._eth_pose = [ws.cell(row=i, column=6).value for i in range(2, self._num_row+1)]
        self._rem_pose = [ws.cell(row=i, column=7).value for i in range(2, self._num_row+1)]
        self._eth_force = [ws.cell(row=i, column=9).value for i in range(2, self._num_row+1)]
        self._rem_force = [ws.cell(row=i, column=10).value for i in range(2, self._num_row+1)]

        self._eth_pose_jerk = [0.0 for i in range(2, self._num_row + 1)]
        self._rem_pose_jerk = [0.0 for i in range(2, self._num_row + 1)]
        self._eth_force_jerk = [0.0 for i in range(2, self._num_row + 1)]
        self._rem_force_jerk = [0.0 for i in range(2, self._num_row + 1)]

    def set_time_start_from_zero(self):
        self._time_hour = [float(self._time_hour[i]) - float(self._time_start[0]) for i in range(0, self._num_row-1)]
        self._time_min = [float(self._time_min[i]) - float(self._time_start[1]) for i in range(0, self._num_row-1)]
        self._time_sec = [float(self._time_sec[i]) - float(self._time_start[2]) for i in range(0, self._num_row-1)]
        self._time = [self._time_hour[i] * 60 * 60 + self._time_min[i] * 60 + self._time_sec[i] for i in range(0, self._num_row - 1)]

    def set_motion_from_start(self):
        self._eth_pose = [self._eth_pose_start - self._eth_pose[i] for i in range(0, self._num_row - 1)]
        self._rem_pose = [self._rem_pose[i] - self._rem_pose_start for i in range(0, self._num_row - 1)]

    def set_force_from_start(self):
        self._eth_force = [self._eth_force[i] - self._eth_force_start for i in range(0, self._num_row - 1)]
        self._rem_force = [self._rem_force[i] - self._rem_force_start for i in range(0, self._num_row - 1)]

    def interaction_time_sec(self):
        if float(self._time_hour[0]) != 0:
            self.set_time_start_from_zero()
            print("set")
        return self._time[self._num_row-2]

    def motion_diff_eth_rem(self):
        if float(self._eth_pose[0]) != 0:
            self.set_motion_from_start()
            print("set")
        return [abs(self._eth_pose_diff[i] - self._rem_pose_diff[i]) for i in range(0, self._num_row - 1)]

    def force_diff_eth_rem(self):
        if float(self._eth_force[0]) != 0:
            self.set_force_from_start()
            print("set")
        return [abs(self._eth_force_diff[i] - self._rem_force_diff[i]) for i in range(0, self._num_row - 1)]

    def motion_jerk(self):
        # Calculate velocity using finite differences
        velocity = np.gradient(self._eth_pose, self._time, edge_order=2)
        acceleration = np.gradient(velocity, self._time, edge_order=2)
        jerk = [abs(np.gradient(acceleration, self._time, edge_order=2))]

        fig, ax = plt.subplots()
        #ax.plot(self._time, self._eth_pose, label='eth motion')
        #ax.plot(self._time, self._rem_pose, label='remote motion')#, marker='o', markersize=2)
        #ax.legend()
        bp = ax.boxplot(self._time, jerk, patch_artist=True)
        plt.show()

        #motion_jerk_rms = np.sqrt(np.mean(jerk ** 2))

        print("motion jerk:", jerk)


    def draw_motion(self):
        if float(self._eth_pose[0]) != 0:
            self.set_motion_from_start()
            print("set")

        fig = plt.figure(figsize=(10, 6))
        ax = plt.axes((0.1, 0.1, 0.5, 0.8))

        ax.plot(self._time, self._eth_pose, label='eth motion')
        ax.plot(self._time, self._rem_pose, label='remote motion')#, marker='o', markersize=2)
        ax.legend()
        plt.show()

    def draw_force(self):
        if float(self._eth_force[0]) != 0:
            self.set_force_from_start()
            print("set")

        fig = plt.figure(figsize=(10, 6))
        ax = plt.axes((0.1, 0.1, 0.5, 0.8))

        ax.plot(self._time, self._eth_force, label='eth force')
        ax.plot(self._time, self._rem_force, label='remote force')#, marker='o', markersize=2)
        ax.legend()
        plt.show()

wb = openpyxl.load_workbook('xlsx/4_all_at_once.xlsx')

Data_20 = Data(wb['20'])
#Data_50 = Data(wb['50'])
#Data_60 = Data(wb['60'])
#Data_Ke = Data(wb['Ke'])

print("Start!")
print(Data_20.interaction_time_sec())
#print(Data_50.interaction_time_sec())
#print(Data_60.interaction_time_sec())
#print(Data_Ke.interaction_time_sec())

# set global parameters
mpl.rcParams['pdf.fonttype'] = 42
mpl.rcParams['ps.fonttype'] = 42
mpl.rcParams['font.family'] = 'Arial'

#Data_20.draw_motion()
#Data_20.draw_force()

Data_20.motion_jerk()



print("End!")
#dataset = [Data_20.set_motion_diff(),Data_50.set_motion_diff(),Data_60.set_motion_diff(), Data_Ke.set_motion_diff()] # the sequence of arrays
#dataset = [Data_20.set_force_diff(),Data_50.set_force_diff(),Data_60.set_force_diff(), Data_Ke.set_force_diff()] # the sequence of arrays
#positions = [1, 3, 5, 7]  # where to put those violin, the x-coordinates

#fig,ax = plt.subplots()
# sequence of arrays # where to put these arrays # allow filling the box with colors
#bp = ax.boxplot(x=dataset, positions=positions, patch_artist=True)
#vp = ax.violinplot(dataset=dataset, positions=positions)
#plt.show()


#Data_20.motion_difference()

#ws = wb['20']
#num_row = ws.max_row

#print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))



# motion profile:
# 1. calculate em = (eth - eth start), rm = (remote - remote start)
# 2. calculate md = abs(em - rm)
# 3. calculate mean/rms/variance of md
# 4. print mean/rms/variance of md
# 5. draw violin plot and box plot of md

#start_eth_p = ws.cell(row=2, column=6).value
#start_rem_p = ws.cell(row=2, column=7).value

#em_i = [ws.cell(row=i, column=6).value - float(start_eth_p) for i in range(2, num_row)]
#rm_i = [ws.cell(row=i, column=7).value - float(start_rem_p) for i in range(2, num_row)]
#md_i = [abs(em_i[i] - rm_i[i]) for i in range(0, num_row-2)]

#data1 = md_i



# force profile:
# 1. calculate ef = eth-3.0, rf = remote-0.1
# 2. calculate fd = abs(ef - rf)
# 3. calculate mean/rms/variance of fd,