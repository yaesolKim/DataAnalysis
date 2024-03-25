import openpyxl

import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.ticker import MultipleLocator, FormatStrFormatter, MaxNLocator

class Data:

    def __init__(self, ws):
        self._num_row = ws.max_row
        self._time = np.array([ws.cell(row=i, column=1).value for i in range(2, self._num_row+1)])

        self._eth_pose = np.array([ws.cell(row=i, column=6).value for i in range(2, self._num_row + 1)])
        self._rem_pose = np.array([ws.cell(row=i, column=7).value for i in range(2, self._num_row + 1)])
        self._eth_force = np.array([ws.cell(row=i, column=9).value for i in range(2, self._num_row + 1)])
        self._rem_force = np.array([ws.cell(row=i, column=10).value for i in range(2, self._num_row + 1)])

        self._eth_pose_jerk = np.array([0.0 for i in range(0, self._num_row - 1)])
        self._rem_pose_jerk = np.array([0.0 for i in range(0, self._num_row - 1)])
        self._eth_force_jerk = np.array([0.0 for i in range(0, self._num_row - 1)])
        self._rem_force_jerk = np.array([0.0 for i in range(0, self._num_row - 1)])

        self._eth_motion_jerk_rms = 0.00
        self._rem_motion_jerk_rms = 0.00
        self._eth_force_jerk_rms = 0.00
        self._rem_force_jerk_rms = 0.00

    def motion_jerk(self):
        self._eth_pose_jerk, self._eth_motion_jerk_rms = calculate_jerk(self._eth_pose, self._time)
        self._rem_pose_jerk, self._rem_motion_jerk_rms = calculate_jerk(self._rem_pose, self._time)

    def force_jerk(self):
        self._eth_force_jerk, self._eth_force_jerk_rms = calculate_jerk(self._eth_force, self._time)
        self._rem_force_jerk, self._rem_force_jerk_rms = calculate_jerk(self._rem_force, self._time)

def calculate_jerk(pose, time):
    velocity = np.gradient(pose, time, edge_order=2)
    acceleration = np.gradient(velocity, time, edge_order=2)
    jerk = np.gradient(acceleration, time, edge_order=2)  # Exclude the first element of time

    jerk_abs = np.abs(jerk)
    jerk_rms = np.sqrt(np.mean(jerk_abs ** 2))
    return jerk_abs, jerk_rms


def plot_jerk(y1, y2, y3, y4, robot, profile):
    fig, axes = plt.subplots(nrows=2, ncols=2, figsize=(10, 6))

    if robot == "eth":
        if profile == "pose":
            y1_y = y1._eth_pose_jerk
            y2_y = y2._eth_pose_jerk
            y3_y = y3._eth_pose_jerk
            y4_y = y4._eth_pose_jerk
            y1_rms = y1._eth_motion_jerk_rms
            y2_rms = y2._eth_motion_jerk_rms
            y3_rms = y3._eth_motion_jerk_rms
            y4_rms = y4._eth_motion_jerk_rms
            y_axis = 6000

        elif profile == "force":
            y1_y = y1._eth_force_jerk
            y2_y = y2._eth_force_jerk
            y3_y = y3._eth_force_jerk
            y4_y = y4._eth_force_jerk
            y1_rms = y1._eth_force_jerk_rms
            y2_rms = y2._eth_force_jerk_rms
            y3_rms = y3._eth_force_jerk_rms
            y4_rms = y4._eth_force_jerk_rms
            y_axis = 700000

    elif robot == "rem":
        if profile == "pose":
            y1_y = y1._rem_pose_jerk
            y2_y = y2._rem_pose_jerk
            y3_y = y3._rem_pose_jerk
            y4_y = y4._rem_pose_jerk
            y1_rms = y1._rem_motion_jerk_rms
            y2_rms = y2._rem_motion_jerk_rms
            y3_rms = y3._rem_motion_jerk_rms
            y4_rms = y4._rem_motion_jerk_rms
            y_axis = 2000

        elif profile == "force":
            y1_y = y1._rem_force_jerk
            y2_y = y2._rem_force_jerk
            y3_y = y3._rem_force_jerk
            y4_y = y4._rem_force_jerk
            y1_rms = y1._rem_force_jerk_rms
            y2_rms = y2._rem_force_jerk_rms
            y3_rms = y3._rem_force_jerk_rms
            y4_rms = y4._rem_force_jerk_rms
            y_axis = 500000

    axes[0, 0].plot(y1._time, y1_y, label='')
    axes[0, 0].set_ylim(0, y_axis)
    axes[0, 0].set_title('K = 20')
    axes[0, 0].text(np.max(y1._time)*0.2, y_axis*0.9, f'Mean Jerk RMS: {y1_rms:.2f}', ha='center', va='bottom')

    axes[0, 1].plot(y2._time, y2_y, label='')
    axes[0, 1].set_ylim(0, y_axis)
    axes[0, 1].set_title('K = 50')
    axes[0, 1].text(np.max(y2._time)*0.2, y_axis*0.9, f'Mean Jerk RMS: {y2_rms:.2f}', ha='center', va='bottom')

    axes[1, 0].plot(y3._time, y3_y, label='')
    axes[1, 0].set_ylim(0, y_axis)
    axes[1, 0].set_title('K = 60')
    axes[1, 0].text(np.max(y3._time)*0.2, y_axis*0.9, f'Mean Jerk RMS: {y3_rms:.2f}', ha='center', va='bottom')

    axes[1, 1].plot(y4._time, y4_y, label='')
    axes[1, 1].set_ylim(0, y_axis)
    axes[1, 1].set_title('K = Ke')
    axes[1, 1].text(np.max(y4._time)*0.2, y_axis*0.9, f'Mean Jerk RMS: {y4_rms:.2f}', ha='center', va='bottom')

    plt.tight_layout()
    plt.show()

wb_name = 'ETH_VIC/data.ETH_VIC'
ws1_name = '20 (2)'
ws2_name = '50 (2)'
ws3_name = '60 (2)'
ws4_name = 'Ke (2)'

wb = openpyxl.load_workbook(wb_name)
Data_20 = Data(wb[ws1_name])
Data_50 = Data(wb[ws2_name])
Data_60 = Data(wb[ws3_name])
Data_Ke = Data(wb[ws4_name])

# set global parameters
mpl.rcParams['pdf.fonttype'] = 42
mpl.rcParams['ps.fonttype'] = 42
mpl.rcParams['font.family'] = 'Arial'

Data_20.motion_jerk()
Data_50.motion_jerk()
Data_60.motion_jerk()
Data_Ke.motion_jerk()

Data_20.force_jerk()
Data_50.force_jerk()
Data_60.force_jerk()
Data_Ke.force_jerk()

plot_jerk(Data_20, Data_50, Data_60, Data_Ke, 'eth', 'pose')
plot_jerk(Data_20, Data_50, Data_60, Data_Ke, 'rem', 'pose')
plot_jerk(Data_20, Data_50, Data_60, Data_Ke, 'eth', 'force')
plot_jerk(Data_20, Data_50, Data_60, Data_Ke, 'rem', 'force')


#plot_jerk(Data_20._time, Data_20._eth_pose_jerk, Data_50._time, Data_50._eth_pose_jerk, Data_60._time, Data_60._eth_pose_jerk, Data_Ke._time, Data_Ke._eth_pose_jerk)
#plot_jerk(Data_20._jerk_time, Data_20._eth_pose_jerk, Data_50._jerk_time, Data_50._eth_pose_jerk, Data_60._jerk_time, Data_60._eth_pose_jerk, Data_Ke._jerk_time, Data_Ke._eth_pose_jerk)
#plot_jerk(Data_20._jerk_time, Data_20._rem_pose_jerk, Data_50._jerk_time, Data_50._rem_pose_jerk, Data_60._jerk_time, Data_60._rem_pose_jerk, Data_Ke._jerk_time, Data_Ke._rem_pose_jerk)
#plot_jerk(Data_20._jerk_time, Data_20._eth_force_jerk, Data_50._jerk_time, Data_50._eth_force_jerk, Data_60._jerk_time, Data_60._eth_force_jerk, Data_Ke._jerk_time, Data_Ke._eth_force_jerk)
#plot_jerk(Data_20._jerk_time, Data_20._rem_force_jerk, Data_50._jerk_time, Data_50._rem_force_jerk, Data_60._jerk_time, Data_60._rem_force_jerk, Data_Ke._jerk_time, Data_Ke._rem_force_jerk)

#draw barplot of task completion time
#bar_plot_time(Data_20.interaction_time_sec(), Data_50.interaction_time_sec(), Data_60.interaction_time_sec(), Data_Ke.interaction_time_sec())

# draw motion profile of each conditions
#plot_motion_raw(Data_20, Data_50, Data_60, Data_Ke) # raw
#plot_motion(Data_20._time, Data_20._eth_pose, Data_20._rem_pose, Data_50._time, Data_50._eth_pose, Data_50._rem_pose, Data_60._time, Data_60._eth_pose, Data_60._rem_pose, Data_Ke._time, Data_Ke._eth_pose, Data_Ke._rem_pose)
#plot_force(Data_20._time, Data_20._eth_force, Data_20._rem_force, Data_50._time, Data_50._eth_force, Data_50._rem_force, Data_60._time, Data_60._eth_force, Data_60._rem_force, Data_Ke._time, Data_Ke._eth_force, Data_Ke._rem_force)