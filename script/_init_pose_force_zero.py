import openpyxl

import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.ticker import MultipleLocator, FormatStrFormatter, MaxNLocator

class Data:

    def __init__(self, ws):
        self._num_row = ws.max_row

        self._time_start = ws.cell(row=2, column=1).value.split('/')[3].split(':')
        self._time = [ws.cell(row=i, column=1).value.split('/')[3] for i in range(2, self._num_row + 1)]

        self._eth_pose = [ws.cell(row=i, column=6).value for i in range(2, self._num_row+1)]
        self._rem_pose = [ws.cell(row=i, column=7).value for i in range(2, self._num_row+1)]
        self._eth_force = [ws.cell(row=i, column=9).value for i in range(2, self._num_row+1)]
        self._rem_force = [ws.cell(row=i, column=10).value for i in range(2, self._num_row+1)]

    def set_time_start_from_zero(self):
        for i in range(0, self._num_row - 1):
            h = float(self._time[i].split(':')[0]) - float(self._time_start[0])
            m = float(self._time[i].split(':')[1]) - float(self._time_start[1])
            s = float(self._time[i].split(':')[2]) - float(self._time_start[2])
            self._time[i] = h*60*60 + m*60 + s

def change_data_in_worksheet(filename, worksheet_name, num_row, new_value):
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook[worksheet_name]

    # Change the value in the specified cell
    for i in range(2, num_row+1):
        worksheet.cell(row=i, column=1, value=new_value[i-2])

    # Save the modified workbook
    workbook.save(filename)
    print(f"Cells updated in worksheet '{worksheet_name}' in '{filename}'.")

wb_name = 'xlsx/data.xlsx'
ws1_name = '20 (2)'
ws2_name = '50 (2)'
ws3_name = '60 (2)'
ws4_name = 'Ke (2)'

wb = openpyxl.load_workbook(wb_name)

Data_20 = Data(wb[ws1_name])
Data_50 = Data(wb[ws2_name])
Data_60 = Data(wb[ws3_name])
Data_Ke = Data(wb[ws4_name])

Data_20.set_time_start_from_zero()
Data_50.set_time_start_from_zero()
Data_60.set_time_start_from_zero()
Data_Ke.set_time_start_from_zero()

change_data_in_worksheet(wb_name, ws1_name, Data_20._num_row, Data_20._time)
change_data_in_worksheet(wb_name, ws2_name, Data_50._num_row, Data_50._time)
change_data_in_worksheet(wb_name, ws3_name, Data_60._num_row, Data_60._time)
change_data_in_worksheet(wb_name, ws4_name, Data_Ke._num_row, Data_Ke._time)
